# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.comments import Comment

F = "Meiryo"
INK   = "1B2330"
ACC   = "1F4E6B"
RULE  = "DFD8CA"
HEADBG= "1F4E6B"
BAND  = "FAF8F3"
INPUT = "FFF3C4"   # 記入欄
GREEN = "1C6647"
AMBER = "8A5A12"
RED   = "993124"
GREY  = "6E747D"

A = "https://claude.ai/code/artifact/"

# 案件, 状態, LP名, デザイン方向, ①, ②, ③, 残っていること, 資料請求, 会社概要, 本編ID(URL), ID, 更新日
rows = [
 ("REALIZE CLUB（法人向け）","現行","会社が、社員の人生を守る時代へ。","深緑 × 和モダン／昭和と令和の対比","○","○","△",
  "送信ボタンがメーラーを開くだけ（info@realize-os.co.jp 宛）。受付フォームの実装が必要／会社概要ページが未作成／本番ドメインへの公開",
  A+"b41aa6c8-3939-483e-9cf4-c92b3290ec3c", "", A+"08b479a1-0803-4978-989b-dfbb853c3e95","08b479a1","2026-09-03"),

 ("REALIZE CLUB（法人向け）","別案","会社が社員の人生を支える時代へ","星空タウン／クリーム × 濃紺のイラスト","○","○","△",
  "RC案で唯一3ページ揃い。採用するかの判断／送信はメーラー起動のみ",
  A+"4e88d497-2710-4ece-9972-60aabc2da022", A+"c5af6a39-a1ec-47e7-a2dd-652458d7fd69", A+"7355f424-d6c3-47fb-88a4-a9cfc890f07e","7355f424","2026-08-31"),

 ("REALIZE CLUB（法人向け）","別案","困る前に、灯す。","漆黒 × 炎のグラデーション","○","△","×",
  "kaisha.html / shiryou-realize.html / sodan.html への相対リンクが残り、単体では飛べない／会社概要ページの作成",
  "（同ファイル内）","", A+"ded9a489-7357-4ed8-80a0-e012da4aaec4","ded9a489","2026-08-31"),

 ("REALIZE CLUB（法人向け）","別案","その安心、今がお得。","スーパーの特売ポップ／白地に赤と黄","○","△","×",
  "リンクがすべてページ内アンカー。外部への導線が1本もない／会社概要ページ、送信先の設定",
  "（同ページ内フォーム）","", A+"4ce48691-4c9a-46fb-bbd2-e375366bc5ab","4ce48691","2026-08-31"),

 ("REALIZE CLUB（法人向け）","別案","社員が、溶けていく前に。","アイスのメタファー／手描き風","○","△","×",
  "会社概要ページが未作成／送信はメーラー起動のみ",
  A+"a631f99f-f0e8-4bc0-8f78-67f01132ed31","", A+"b0f16478-7326-4826-88ce-4aa0230fbcae","b0f16478","2026-08-28"),

 ("LIFE MAKE PARTNERS（加盟店募集）","現行","家より先に、人生の話を。","韓国料理店の看板／木の壁 × 生成りの横断幕","○","○","○",
  "本番ドメインへの公開だけが残り。問い合わせはLINE公式（lin.ee/WHYApuM）に集約済み",
  A+"00bc4beb-3bef-4d45-a94b-90c693a3c746", A+"5d670aad-cdef-49f7-bf67-38b65e8c396e", A+"7f212d2f-1999-4a55-ba64-865fbcc1b6bf","7f212d2f","2026-09-03"),

 ("LIFE MAKE PARTNERS（加盟店募集）","別案","いつも同じ集客ばっかやってんじゃねぇ！","喝・挑発ポスター／赤と黄の斜めストライプ","○","△","×",
  "kaisha-katsu.html / shiryou-katsu.html への相対リンクが残り飛べない／行き先が「#」のままのボタンが2つ",
  "（同ページ内フォーム）","", A+"db82fc95-4c62-4b83-b22c-e7cee86bc870","db82fc95","2026-08-28"),

 ("LIFE MAKE PARTNERS（加盟店募集）","別案","加盟資格は、人生への愛。ただそれだけ。","星空タウン／RC星空案と同じ世界観","○","○","△",
  "行き先が「#」のままのボタンが2つ／現行案との統廃合の判断",
  A+"32b51fff-748e-4fb3-a257-99531381bd35", A+"cb3eacf9-1c9d-4722-a6d0-a599ed2fe3f5", A+"65b43069-9f73-4f01-881f-70c31cd5cc30","65b43069","2026-08-24"),

 ("まちの相談窓口","現行","不動産屋が、まちの相談窓口になる。","銭湯ポスター（リソグラフ2色刷り）／LINE誘導版","○","○","△",
  "行き先が「#」のままのリンクが2つ残っている／本番ドメインへの公開",
  A+"dec5629a-75b7-4a8d-ba24-a6e1b428c0d3", A+"0d2281f1-4a5f-4592-a013-1b5710b515b4", A+"1a259364-2cf9-4098-b864-bde6880ad6bf","1a259364","2026-09-03"),

 ("まちの相談窓口","旧版","まちの相談窓口（フォーム版）","銭湯ポスター／申込フォーム内蔵","○","○","△",
  "現行のLINE誘導版に置き換わっている。残すか消すかの判断",
  A+"dec5629a-75b7-4a8d-ba24-a6e1b428c0d3", A+"0d2281f1-4a5f-4592-a013-1b5710b515b4", A+"39f0d878-53ee-4d73-9625-e9ce060b7aa8","39f0d878","2026-08-20"),

 ("不動産FC加盟","単発","収益が積み上がる不動産FC加盟へ。","ボクシング／黒地 × 赤とゴールド","○","×","×",
  "ページのタイトルタグが空。共有すると名前なしで出る／資料請求・会社概要ページ、送信先の設定",
  "（同ページ内フォーム）","", A+"d5160792-9b05-48a1-aa52-e88cff874321","d5160792","2026-08-19"),
]

HEAD = ["案件","状態","LP名","デザインの方向","①原稿・デザイン","②付随ページ",
        "③問い合わせ導線","④本番公開","到達度","本番URL（記入欄）","残っていること",
        "資料請求ページ","会社概要ページ","LPプレビュー","ID","更新日"]
W = [22,7,30,32,13,12,14,11,9,34,52,30,30,30,10,11]

wb = Workbook()

# ================= LP一覧 =================
ws = wb.active
ws.title = "LP一覧"

thin = Side(style="thin", color=RULE)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "LP制作台帳"
ws["A1"].font = Font(name=F, size=16, bold=True, color=INK)
ws["A2"] = "これまでに作った11本のLP。何を作ったか／どこまでできているか／公開後の本番URLを1枚で管理する。調査日 2026-09-03"
ws["A2"].font = Font(name=F, size=9, color=GREY)
ws["A3"] = "黄色のJ列だけ記入してください。URLを入れると④本番公開と到達度が自動で変わります。"
ws["A3"].font = Font(name=F, size=9, bold=True, color=AMBER)

HR = 5
for i, h in enumerate(HEAD, start=1):
    c = ws.cell(row=HR, column=i, value=h)
    c.font = Font(name=F, size=9, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=HEADBG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
ws.row_dimensions[HR].height = 34

for i, w in enumerate(W, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

def link(cell, url, label):
    cell.value = label
    if url.startswith("http"):
        cell.hyperlink = url
        cell.font = Font(name=F, size=9, color=ACC, underline="single")
    else:
        cell.font = Font(name=F, size=9, color=GREY)

for n, r in enumerate(rows):
    ex = HR + 1 + n
    (anken, state, name, direction, s1, s2, s3, todo, doc, comp, prev, aid, upd) = r

    ws.cell(row=ex, column=1, value=anken)
    ws.cell(row=ex, column=2, value=state)
    ws.cell(row=ex, column=3, value=name)
    ws.cell(row=ex, column=4, value=direction)
    ws.cell(row=ex, column=5, value=s1)
    ws.cell(row=ex, column=6, value=s2)
    ws.cell(row=ex, column=7, value=s3)
    # ④本番公開 = J列(本番URL)が入っていれば○
    ws.cell(row=ex, column=8, value='=IF(J{0}="","×","○")'.format(ex))
    # 到達度 = ○を1、△を0.5として4段のうち何割か
    ws.cell(row=ex, column=9,
            value='=COUNTIF(E{0}:H{0},"○")+COUNTIF(E{0}:H{0},"△")'.format(ex))
    ws.cell(row=ex, column=10, value=None)          # 本番URL（記入欄）
    ws.cell(row=ex, column=11, value=todo)
    link(ws.cell(row=ex, column=12), doc,  doc  if not doc.startswith("http") else "資料請求を開く")
    link(ws.cell(row=ex, column=13), comp, "会社概要を開く" if comp.startswith("http") else "未作成")
    link(ws.cell(row=ex, column=14), prev, "LPを開く")
    ws.cell(row=ex, column=15, value=aid)
    ws.cell(row=ex, column=16, value=upd)

    ws.row_dimensions[ex].height = 46
    for col in range(1, len(HEAD) + 1):
        c = ws.cell(row=ex, column=col)
        c.border = box
        if col not in (12, 13, 14):
            c.font = Font(name=F, size=9, color=INK)
        c.alignment = Alignment(vertical="center", wrap_text=(col in (1,3,4,11)),
                                horizontal="center" if col in (2,5,6,7,8,9,16) else "left")
        if n % 2 == 1 and col != 10:
            c.fill = PatternFill("solid", fgColor=BAND)

    ws.cell(row=ex, column=2).font = Font(name=F, size=9, bold=True,
        color={"現行":GREEN,"別案":AMBER,"旧版":GREY,"単発":GREY}[state])
    ws.cell(row=ex, column=9).number_format = '0"／4段"' 
    # 記入欄
    j = ws.cell(row=ex, column=10)
    j.fill = PatternFill("solid", fgColor=INPUT)
    j.font = Font(name=F, size=9, color="0000FF")
    j.alignment = Alignment(vertical="center", horizontal="left")

first, last = HR + 1, HR + len(rows)
ws.cell(row=HR+1, column=10).comment = Comment(
    "公開した本番LPのURLをここに貼ってください。\n入力すると④本番公開が○になり、到達度が上がります。", "LP制作台帳", height=90, width=260)

# 未着手の段を赤く
ws.conditional_formatting.add("E{0}:H{1}".format(first, last),
    CellIsRule(operator="equal", formula=['"×"'], font=Font(name=F, size=9, bold=True, color=RED)))
ws.conditional_formatting.add("E{0}:H{1}".format(first, last),
    CellIsRule(operator="equal", formula=['"△"'], font=Font(name=F, size=9, bold=True, color=AMBER)))
ws.conditional_formatting.add("E{0}:H{1}".format(first, last),
    CellIsRule(operator="equal", formula=['"○"'], font=Font(name=F, size=9, bold=True, color=GREEN)))

dv = DataValidation(type="list", formula1='"○,△,×"', allow_blank=False)
ws.add_data_validation(dv)
dv.add("E{0}:G{1}".format(first, last))

# 合計行
tot = last + 1
ws.cell(row=tot, column=4, value="合計 / 平均").font = Font(name=F, size=9, bold=True, color=INK)
ws.cell(row=tot, column=4).alignment = Alignment(horizontal="right")
for col, f in ((5,'=COUNTIF(E{0}:E{1},"○")'), (6,'=COUNTIF(F{0}:F{1},"○")'),
               (7,'=COUNTIF(G{0}:G{1},"○")'), (8,'=COUNTIF(H{0}:H{1},"○")')):
    c = ws.cell(row=tot, column=col, value=f.format(first, last))
    c.font = Font(name=F, size=9, bold=True, color=ACC)
    c.alignment = Alignment(horizontal="center")
    c.number_format = '0"／11"'
c = ws.cell(row=tot, column=9, value="=AVERAGE(I{0}:I{1})".format(first, last))
c.font = Font(name=F, size=9, bold=True, color=ACC)
c.number_format = '0.0"／4段"' 
c.alignment = Alignment(horizontal="center")
c = ws.cell(row=tot, column=10, value='=COUNTA(J{0}:J{1})&" 本 登録済み"'.format(first, last))
c.font = Font(name=F, size=9, bold=True, color=ACC)
for col in range(4, 11):
    ws.cell(row=tot, column=col).border = Border(top=Side(style="medium", color=ACC))

ws.freeze_panes = "D6"
ws.auto_filter.ref = "A{0}:P{1}".format(HR, last)
ws.sheet_view.showGridLines = False

# ================= 凡例 =================
lg = wb.create_sheet("凡例と調べ方")
lg.sheet_view.showGridLines = False
lg.column_dimensions["A"].width = 22
lg.column_dimensions["B"].width = 96

def head(r, t):
    c = lg.cell(row=r, column=1, value=t)
    c.font = Font(name=F, size=12, bold=True, color=ACC)

def kv(r, k, v):
    a = lg.cell(row=r, column=1, value=k); a.font = Font(name=F, size=9, bold=True, color=INK)
    a.alignment = Alignment(vertical="top")
    b = lg.cell(row=r, column=2, value=v); b.font = Font(name=F, size=9, color=INK)
    b.alignment = Alignment(vertical="top", wrap_text=True)
    lg.row_dimensions[r].height = 30

lg["A1"] = "凡例と、この表の作り方"
lg["A1"].font = Font(name=F, size=15, bold=True, color=INK)

head(3, "「どこまで」の4段")
kv(4, "①原稿・デザイン", "本編の文章とデザインが最後まで通っている")
kv(5, "②付随ページ", "資料請求・会社概要など、必要な付随ページが揃っている")
kv(6, "③問い合わせ導線", "ボタンの行き先が実在し、問い合わせが届く先につながっている")
kv(7, "④本番公開", "独自ドメインで公開され、J列に本番URLが入っている（自動判定）")

head(9, "記号")
kv(10, "○", "できている")
kv(11, "△", "仮づけ。メーラーを開くだけ、または行き先が「#」のまま")
kv(12, "×", "手つかず／ページが存在しない")

head(14, "状態")
kv(15, "現行", "その案件で採用している案")
kv(16, "別案", "同じ商材でデザインを変えて作った比較用の案")
kv(17, "旧版", "新しい版に置き換わったもの")
kv(18, "単発", "他と系統がつながっていない単発のLP")

head(20, "使い方")
kv(21, "J列（黄色）", "公開した本番LPのURLを貼ります。ここだけが記入欄です。入力すると④本番公開が○に変わり、到達度と合計行が自動で更新されます。")
kv(22, "到達度", "①〜④のうち、○か△が付いている段の数を数えています。△は「到達しているが仮づけ」なので段数には入ります。数式なので手で書き換えないでください。")
kv(23, "E〜G列", "作業が進んだら ○ △ × を選び直せます（プルダウン）。")

head(25, "この表の作り方")
kv(26, "調査範囲", "公開済みアーティファクト43件のうち、LPに該当する22ページ（LP本編11本＋付随ページ11本）。")
kv(27, "確認したこと", "1ページずつHTMLを読み、ボタンの行き先・フォームの送信方法・付随ページの有無を実物で確認しました。「残っていること」は確認できた事実だけを書いています。")
kv(28, "未確認", "付随ページ4本（RC星空案の会社概要、RCアイス案の資料請求、LMP星空案の資料請求と会社概要）は、親LPからのリンクで存在を確認しただけで、中身は個別に開いていません。")
kv(29, "調査日", "2026-09-03")

for r in (4,5,6,7,10,11,12,15,16,17,18,21,22,23,26,27,28,29):
    lg.row_dimensions[r].height = 34

# openpyxl はキャッシュ値を書かないので、Excel で開いた瞬間に全式を計算させる
wb.calculation.fullCalcOnLoad = True

wb.save("/home/user/jigyokeikakusyo/LP制作台帳.xlsx")
print("saved")
