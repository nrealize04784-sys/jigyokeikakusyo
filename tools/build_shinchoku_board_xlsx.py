# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.comments import Comment

F="Meiryo"
INK="1B2330"; ACC="1F4E6B"; RULE="DFD8CA"; BAND="FAF8F3"; INPUT="FFF3C4"
GREEN="1C6647"; AMBER="8A5A12"; RED="993124"; GREY="6E747D"
EDIT="EAF1F6"   # 選んで変えられるセル
A="https://claude.ai/code/artifact/"
thin=Side(style="thin",color=RULE); box=Border(left=thin,right=thin,top=thin,bottom=thin)

wb=Workbook()

# 進捗のドロップダウン用の値を、非表示シートに置く。
# 範囲を参照する入力規則にすると、Excelのドロップダウンに「20%」と書式つきで並ぶ。
opt=wb.create_sheet("選択肢")
opt["A1"]="進捗"; opt["A1"].font=Font(name=F,size=9,bold=True)
for i in range(21):
    c=opt.cell(row=2+i,column=1,value=i*0.05)
    c.number_format="0%"; c.font=Font(name=F,size=9)

ANKEN=["REALIZE CLUB（法人向け）","LIFE MAKE PARTNERS（加盟店募集）","まちの相談窓口","不動産FC加盟"]
LPSTATE=["現行","別案","旧版","単発"]
KUBUN=["LP","教材・ツール","動画","広告"]
PRIORITY=["P1","P2","P3"]
STATES=["未着手","進行中","確認待ち"]
for col,(title,vals) in enumerate(
        [("案件",ANKEN),("LPの状態",LPSTATE),("区分",KUBUN),("優先",PRIORITY),("状態",STATES)], start=2):
    h=opt.cell(row=1,column=col,value=title); h.font=Font(name=F,size=9,bold=True)
    opt.column_dimensions[get_column_letter(col)].width=30
    for i,v in enumerate(vals):
        opt.cell(row=2+i,column=col,value=v).font=Font(name=F,size=9)

def rng(col,n): return "'選択肢'!${0}$2:${0}${1}".format(col,n+1)
PCT_LIST   = "'選択肢'!$A$2:$A$22"
ANKEN_LIST = rng("B",len(ANKEN))
LPST_LIST  = rng("C",len(LPSTATE))
KUBUN_LIST = rng("D",len(KUBUN))
PRI_LIST   = rng("E",len(PRIORITY))
STATE_LIST = rng("F",len(STATES))

def sheet_header(ws, title, lead, note):
    ws["A1"]=title; ws["A1"].font=Font(name=F,size=16,bold=True,color=INK)
    ws["A2"]=lead;  ws["A2"].font=Font(name=F,size=9,color=GREY)
    ws["A3"]=note;  ws["A3"].font=Font(name=F,size=9,bold=True,color=AMBER)

def write_head(ws,row,head,widths):
    for i,h in enumerate(head,1):
        c=ws.cell(row=row,column=i,value=h)
        c.font=Font(name=F,size=9,bold=True,color="FFFFFF")
        c.fill=PatternFill("solid",fgColor=ACC)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=box
    ws.row_dimensions[row].height=34
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

# ============================================================ 1. LP
ws=wb.active; ws.title="LP"
HEAD=["案件","状態","LP名","デザインの方向","①原稿・デザイン","②付随ページ","③問い合わせ導線",
      "④本番公開","到達度","本番URL（記入欄）","残っていること","資料請求ページ","会社概要ページ","LPプレビュー","ID","更新日"]
W=[22,7,30,32,13,12,14,11,10,34,52,26,26,26,10,11]
sheet_header(ws,"制作進捗ボード ─ LP",
  "これまでに作った11本のLP。実物を1ページずつ読んで確認したもの。調査日 2026-09-03",
  "黄色のJ列だけ記入してください。URLを入れると④本番公開と到達度が自動で変わります。")
HR=5; write_head(ws,HR,HEAD,W)

lp=[
("REALIZE CLUB（法人向け）","現行","会社が、社員の人生を守る時代へ。","深緑 × 和モダン／昭和と令和の対比","○","○","△",
 "送信ボタンがメーラーを開くだけ。受付フォームの実装が必要／会社概要ページが未作成／本番ドメインへの公開",
 A+"b41aa6c8-3939-483e-9cf4-c92b3290ec3c","",A+"08b479a1-0803-4978-989b-dfbb853c3e95","08b479a1","2026-09-03"),
("REALIZE CLUB（法人向け）","別案","会社が社員の人生を支える時代へ","星空タウン／クリーム × 濃紺のイラスト","○","○","△",
 "RC案で唯一3ページ揃い。採用するかの判断／送信はメーラー起動のみ",
 A+"4e88d497-2710-4ece-9972-60aabc2da022",A+"c5af6a39-a1ec-47e7-a2dd-652458d7fd69",A+"7355f424-d6c3-47fb-88a4-a9cfc890f07e","7355f424","2026-08-31"),
("REALIZE CLUB（法人向け）","別案","困る前に、灯す。","漆黒 × 炎のグラデーション","○","△","×",
 "kaisha.html ほかへの相対リンクが残り、単体では飛べない／会社概要ページの作成",
 "（同ファイル内）","",A+"ded9a489-7357-4ed8-80a0-e012da4aaec4","ded9a489","2026-08-31"),
("REALIZE CLUB（法人向け）","別案","その安心、今がお得。","スーパーの特売ポップ／白地に赤と黄","○","△","×",
 "リンクがすべてページ内アンカー。外部への導線が1本もない／会社概要ページ、送信先の設定",
 "（同ページ内フォーム）","",A+"4ce48691-4c9a-46fb-bbd2-e375366bc5ab","4ce48691","2026-08-31"),
("REALIZE CLUB（法人向け）","別案","社員が、溶けていく前に。","アイスのメタファー／手描き風","○","△","×",
 "会社概要ページが未作成／送信はメーラー起動のみ",
 A+"a631f99f-f0e8-4bc0-8f78-67f01132ed31","",A+"b0f16478-7326-4826-88ce-4aa0230fbcae","b0f16478","2026-08-28"),
("LIFE MAKE PARTNERS（加盟店募集）","現行","家より先に、人生の話を。","韓国料理店の看板／木の壁 × 生成りの横断幕","○","○","○",
 "本番ドメインへの公開だけが残り。問い合わせはLINE公式（lin.ee/WHYApuM）に集約済み",
 A+"00bc4beb-3bef-4d45-a94b-90c693a3c746",A+"5d670aad-cdef-49f7-bf67-38b65e8c396e",A+"7f212d2f-1999-4a55-ba64-865fbcc1b6bf","7f212d2f","2026-09-03"),
("LIFE MAKE PARTNERS（加盟店募集）","別案","いつも同じ集客ばっかやってんじゃねぇ！","喝・挑発ポスター／赤と黄の斜めストライプ","○","△","×",
 "kaisha-katsu.html ほかへの相対リンクが残り飛べない／行き先が「#」のままのボタンが2つ",
 "（同ページ内フォーム）","",A+"db82fc95-4c62-4b83-b22c-e7cee86bc870","db82fc95","2026-08-28"),
("LIFE MAKE PARTNERS（加盟店募集）","別案","加盟資格は、人生への愛。ただそれだけ。","星空タウン／RC星空案と同じ世界観","○","○","△",
 "行き先が「#」のままのボタンが2つ／現行案との統廃合の判断",
 A+"32b51fff-748e-4fb3-a257-99531381bd35",A+"cb3eacf9-1c9d-4722-a6d0-a599ed2fe3f5",A+"65b43069-9f73-4f01-881f-70c31cd5cc30","65b43069","2026-08-24"),
("まちの相談窓口","現行","不動産屋が、まちの相談窓口になる。","銭湯ポスター（リソグラフ2色刷り）／LINE誘導版","○","○","△",
 "行き先が「#」のままのリンクが2つ残っている／本番ドメインへの公開",
 A+"dec5629a-75b7-4a8d-ba24-a6e1b428c0d3",A+"0d2281f1-4a5f-4592-a013-1b5710b515b4",A+"1a259364-2cf9-4098-b864-bde6880ad6bf","1a259364","2026-09-03"),
("まちの相談窓口","旧版","まちの相談窓口（フォーム版）","銭湯ポスター／申込フォーム内蔵","○","○","△",
 "現行のLINE誘導版に置き換わっている。残すか消すかの判断",
 A+"dec5629a-75b7-4a8d-ba24-a6e1b428c0d3",A+"0d2281f1-4a5f-4592-a013-1b5710b515b4",A+"39f0d878-53ee-4d73-9625-e9ce060b7aa8","39f0d878","2026-08-20"),
("不動産FC加盟","単発","収益が積み上がる不動産FC加盟へ。","ボクシング／黒地 × 赤とゴールド","○","×","×",
 "ページのタイトルタグが空。共有すると名前なしで出る／資料請求・会社概要ページ、送信先の設定",
 "（同ページ内フォーム）","",A+"d5160792-9b05-48a1-aa52-e88cff874321","d5160792","2026-08-19"),
]

def link(cell,url,label):
    cell.value=label
    if url.startswith("http"):
        cell.hyperlink=url; cell.font=Font(name=F,size=9,color=ACC,underline="single")
    else:
        cell.font=Font(name=F,size=9,color=GREY)

for n,r in enumerate(lp):
    ex=HR+1+n
    anken,state,name,d,s1,s2,s3,todo,doc,comp,prev,aid,upd=r
    for col,val in ((1,anken),(2,state),(3,name),(4,d),(5,s1),(6,s2),(7,s3),(11,todo),(15,aid),(16,upd)):
        ws.cell(row=ex,column=col,value=val)
    ws.cell(row=ex,column=8,value='=IF(J{0}="","×","○")'.format(ex))
    ws.cell(row=ex,column=9,value='=COUNTIF(E{0}:H{0},"○")+COUNTIF(E{0}:H{0},"△")'.format(ex))
    link(ws.cell(row=ex,column=12),doc,"資料請求を開く" if doc.startswith("http") else doc)
    link(ws.cell(row=ex,column=13),comp,"会社概要を開く" if comp.startswith("http") else "未作成")
    link(ws.cell(row=ex,column=14),prev,"LPを開く")
    ws.row_dimensions[ex].height=46
    for col in range(1,len(HEAD)+1):
        c=ws.cell(row=ex,column=col); c.border=box
        if col not in (12,13,14): c.font=Font(name=F,size=9,color=INK)
        c.alignment=Alignment(vertical="center",wrap_text=(col in (1,3,4,11)),
                              horizontal="center" if col in (2,5,6,7,8,9,16) else "left")
        if n%2==1 and col!=10: c.fill=PatternFill("solid",fgColor=BAND)
    ws.cell(row=ex,column=2).font=Font(name=F,size=9,bold=True,
        color={"現行":GREEN,"別案":AMBER,"旧版":GREY,"単発":GREY}[state])
    ws.cell(row=ex,column=9).number_format='0"／4段"'
    j=ws.cell(row=ex,column=10)
    j.fill=PatternFill("solid",fgColor=INPUT); j.font=Font(name=F,size=9,color="0000FF")
    j.alignment=Alignment(vertical="center")

first,last=HR+1,HR+len(lp)
ws.cell(row=HR+1,column=10).comment=Comment(
 "公開した本番LPのURLをここに貼ってください。\n入力すると④本番公開が○になり、到達度が上がります。","制作進捗ボード",height=90,width=260)
for mark,color in (("×",RED),("△",AMBER),("○",GREEN)):
    ws.conditional_formatting.add("E{0}:H{1}".format(first,last),
        CellIsRule(operator="equal",formula=['"%s"'%mark],font=Font(name=F,size=9,bold=True,color=color)))
dv=DataValidation(type="list",formula1='"○,△,×"',allow_blank=False)
dv.prompt="○＝できている／△＝仮づけ／×＝手つかず"; dv.promptTitle="選んでください"
ws.add_data_validation(dv); dv.add("E{0}:G{1}".format(first,last))
dva=DataValidation(type="list",formula1=ANKEN_LIST,allow_blank=False)
dva.prompt="どの案件のLPかを選びます"; dva.promptTitle="案件を選ぶ"
ws.add_data_validation(dva); dva.add("A{0}:A{1}".format(first,last))
dvb=DataValidation(type="list",formula1=LPST_LIST,allow_blank=False)
dvb.prompt="現行／別案／旧版／単発"; dvb.promptTitle="状態を選ぶ"
ws.add_data_validation(dvb); dvb.add("B{0}:B{1}".format(first,last))
for col in "ABEFG":
    for r in range(first,last+1):
        ws["{0}{1}".format(col,r)].fill=PatternFill("solid",fgColor=EDIT)
ws.conditional_formatting.add("I{0}:I{1}".format(first,last),
    DataBarRule(start_type="num",start_value=0,end_type="num",end_value=4,
                color=ACC,showValue=True))

tot=last+1
c=ws.cell(row=tot,column=4,value="合計 / 平均"); c.font=Font(name=F,size=9,bold=True,color=INK)
c.alignment=Alignment(horizontal="right")
for col in (5,6,7,8):
    L=get_column_letter(col)
    c=ws.cell(row=tot,column=col,value='=COUNTIF({0}{1}:{0}{2},"○")'.format(L,first,last))
    c.font=Font(name=F,size=9,bold=True,color=ACC); c.alignment=Alignment(horizontal="center")
    c.number_format='0"／11"'
c=ws.cell(row=tot,column=9,value="=AVERAGE(I{0}:I{1})".format(first,last))
c.font=Font(name=F,size=9,bold=True,color=ACC); c.number_format='0.0"／4段"'
c.alignment=Alignment(horizontal="center")
c=ws.cell(row=tot,column=10,value='=COUNTA(J{0}:J{1})&" 本 公開済み"'.format(first,last))
c.font=Font(name=F,size=9,bold=True,color=ACC)
for col in range(4,11):
    ws.cell(row=tot,column=col).border=Border(top=Side(style="medium",color=ACC))
ws.freeze_panes="D6"; ws.auto_filter.ref="A{0}:P{1}".format(HR,last); ws.sheet_view.showGridLines=False

# ============================================================ 共通：タスク表
THEAD=["区分","優先","何を作るか／するか","どこで使う","担当","進捗","状態","URL記入欄","メモ"]
TW=[16,6,34,30,22,9,10,34,44]

def task_sheet(name,title,lead,note,groups):
    s=wb.create_sheet(name)
    sheet_header(s,title,lead,note)
    write_head(s,5,THEAD,TW)
    r=6; inputs=[]
    for g,items in groups:
        for it in items:
            kubun,pri,nm,where,own,pct,state,memo=it
            for col,val in ((1,g),(2,pri),(3,nm),(4,where),(5,own),(6,pct/100.0),(7,state),(9,memo)):
                s.cell(row=r,column=col,value=val)
            s.row_dimensions[r].height=32
            for col in range(1,len(THEAD)+1):
                c=s.cell(row=r,column=col); c.border=box; c.font=Font(name=F,size=9,color=INK)
                c.alignment=Alignment(vertical="center",wrap_text=(col in (3,4,9)),
                                      horizontal="center" if col in (2,6,7) else "left")
            s.cell(row=r,column=6).number_format="0%"
            s.cell(row=r,column=7).font=Font(name=F,size=9,bold=True,
                color={"進行中":AMBER,"確認待ち":GREEN,"未着手":GREY}[state])
            if own=="未定":
                s.cell(row=r,column=5).font=Font(name=F,size=9,bold=True,color=RED)
            j=s.cell(row=r,column=8)
            j.fill=PatternFill("solid",fgColor=INPUT); j.font=Font(name=F,size=9,color="0000FF")
            inputs.append(r); r+=1
    f,l=6,r-1
    # 進捗：セルの中にバーを出す
    s.conditional_formatting.add("F{0}:F{1}".format(f,l),
        DataBarRule(start_type="num",start_value=0,end_type="num",end_value=1,
                    color=ACC,showValue=True))
    # 進捗：クリックすると 0%〜100% が並ぶドロップダウン
    dvp=DataValidation(type="list",formula1=PCT_LIST,allow_blank=True)
    dvp.prompt="5%きざみで選べます。直接入力もできます。"; dvp.promptTitle="進捗を選ぶ"
    s.add_data_validation(dvp); dvp.add("F{0}:F{1}".format(f,l))
    # 区分：ドロップダウン（LP／教材・ツール／動画／広告）
    dvk=DataValidation(type="list",formula1=KUBUN_LIST,allow_blank=False)
    dvk.prompt="LP／教材・ツール／動画／広告"; dvk.promptTitle="区分を選ぶ"
    s.add_data_validation(dvk); dvk.add("A{0}:A{1}".format(f,l))
    # 優先：ドロップダウン
    dvpr=DataValidation(type="list",formula1=PRI_LIST,allow_blank=False)
    dvpr.prompt="工程表の優先度"; dvpr.promptTitle="優先を選ぶ"
    s.add_data_validation(dvpr); dvpr.add("B{0}:B{1}".format(f,l))
    # 状態：ドロップダウン
    dvs=DataValidation(type="list",formula1=STATE_LIST,allow_blank=False)
    dvs.prompt="未着手／進行中／確認待ち"; dvs.promptTitle="状態を選ぶ"
    s.add_data_validation(dvs); dvs.add("G{0}:G{1}".format(f,l))
    for rr in range(f,l+1):
        for cl in ("A","B","F","G"):
            s["{0}{1}".format(cl,rr)].fill=PatternFill("solid",fgColor=EDIT)
    t=r
    c=s.cell(row=t,column=5,value="平均"); c.font=Font(name=F,size=9,bold=True,color=INK)
    c.alignment=Alignment(horizontal="right")
    c=s.cell(row=t,column=6,value="=AVERAGE(F{0}:F{1})".format(f,l))
    c.font=Font(name=F,size=9,bold=True,color=ACC); c.number_format="0%"
    c.alignment=Alignment(horizontal="center")
    c=s.cell(row=t,column=7,value='=COUNTIF(G{0}:G{1},"未着手")&" 件未着手"'.format(f,l))
    c.font=Font(name=F,size=9,bold=True,color=ACC)
    c=s.cell(row=t,column=8,value='=COUNTA(H{0}:H{1})&" 件 登録済み"'.format(f,l))
    c.font=Font(name=F,size=9,bold=True,color=ACC)
    for col in range(5,9):
        s.cell(row=t,column=col).border=Border(top=Side(style="medium",color=ACC))
    s.cell(row=f,column=8).comment=Comment(
      "できあがったものの URL を貼ってください。\n動画なら公開URL、広告なら入稿先やアカウントのURL。","制作進捗ボード",height=80,width=250)
    s.freeze_panes="C6"; s.auto_filter.ref="A5:I{0}".format(l); s.sheet_view.showGridLines=False
    return s

# ============================================================ 2. eラーニング・動画
edu=[("教材・ツール",[
 ("eラーニング","P2","eラーニング（ツール本体）","LMP事業 · Web・ツール・システム","社長・白石・片山",50,"進行中","加盟店が学ぶ画面そのもの"),
 ("eラーニング","P2","eラーニング動画","LMP事業 · Web・ツール・システム","未定",20,"進行中","教材の中で流す動画。担当が決まっていない"),
 ("eラーニング","P2","教育診断","RC事業 · Web・ツール・システム","白石",80,"進行中","学ぶ前に現在地をはかる"),
 ("eラーニング","P2","教科書の前段階の資料","RC事業 · 動画・コンテンツ","白石",60,"進行中","教材のもとになる原稿"),
 ("eラーニング","P2","各ツールマニュアル作成","LMP事業 · 営業・資料","未定",0,"未着手","担当が決まっていない"),
 ("eラーニング","P2","研修会準備","LMP事業 · 加盟開発","白石・池之上・大熊",0,"未着手",""),
 ("eラーニング","P2","eラーニング接続・設計","RCP事業 · Web・ツール・システム","白石",0,"未着手",""),
 ("eラーニング","P2","AIロープレ（アオ先生）","AIアオ先生 · 営業コーチ","池之上",90,"確認待ち","制作MASTERのみに記載"),
]),("動画",[
 ("動画","P2","結婚後の暮らしと人生設計動画","RC事業 · 動画・コンテンツ","白石",70,"進行中","動画で唯一動いている1本"),
 ("動画","P2","サービス説明動画","RC事業 · 動画・コンテンツ","白石",0,"未着手",""),
 ("動画","P2","toC動画","RC事業 · 動画・コンテンツ","白石",0,"未着手",""),
 ("動画","P2","サービスデモ動画","RCP事業 · 動画・コンテンツ","白石",0,"未着手",""),
 ("動画","P2","水谷先生の本の動画","RCP事業 · 動画・コンテンツ","白石",0,"未着手",""),
 ("動画","P3","OEM各先生の宣伝動画（5本）","RCP PPモデル · OEM","白石",0,"未着手","AKさん・大橋会長・木原さん・水谷先生・藤田先生"),
 ("動画","P3","宣伝動画・PV ロング／ショート（7本）","RC／RCP／HI／PP · REALIZE QUEST","白石・片山",0,"未着手",""),
 ("動画","P3","採用まわりの動画（2本）","採用 · 学生募集","白石・片山",0,"未着手","PR動画・SNS採用募集動画"),
])]
s2=task_sheet("eラーニング・動画","制作進捗ボード ─ eラーニング・動画",
  "工程表（制作MASTER 2026-09-03 ／ 仕事の棚卸し 2026-09-02）の進捗をそのまま引いています。",
  "ツールの箱は50%まで来ていますが、中に入れる動画が20%で止まっています。赤い担当名は「未定」です。",
  edu)
# 成果物
r=s2.max_row+2
s2.cell(row=r,column=1,value="できている設計書・ガイド").font=Font(name=F,size=11,bold=True,color=ACC)
r+=1
for nm,kind,upd,url in [
 ("LIFE CHECK42 Eラーニング画面設計書","画面設計","2026-08-31",A+"1ce7243d-ad99-47d0-b0d0-ad85c8cc7f16"),
 ("RCP ノーマル版 eラーニング｜UI提案 v1","UI提案","2026-08-05",A+"0ebdd065-594c-447b-a2d6-c1eb4c80d933"),
 ("AIロープレ 操作ガイド","操作ガイド","2026-08-31",A+"fc25e6a1-4589-43ce-8f16-5b416e3fc619"),
 ("生成AI動画講座 比較検討","比較検討","2026-08-21",A+"98260a43-003c-4653-af3c-d667139ea6f8"),
]:
    s2.cell(row=r,column=3,value=nm).font=Font(name=F,size=9,color=INK)
    s2.cell(row=r,column=4,value=kind).font=Font(name=F,size=9,color=GREY)
    s2.cell(row=r,column=5,value=upd).font=Font(name=F,size=9,color=GREY)
    c=s2.cell(row=r,column=9,value="開く"); c.hyperlink=url
    c.font=Font(name=F,size=9,color=ACC,underline="single")
    r+=1

# ============================================================ 3. 広告
ad=[("広告",[
 ("広告","P2","広告制作","RC法人開拓 · 地域企業への入口","大熊・片山",0,"未着手","バナー・原稿などのクリエイティブ"),
 ("広告","P2","広告運用","RC法人開拓／LMP加盟開発","池之上・片山",0,"未着手","出稿・予算・改善"),
 ("広告","P2","広告クリエイティブチェック","LMP事業 · 全体品質確認","白石",0,"未着手","出す前の品質確認"),
 ("広告","P3","広告表示のチェック（景表法）","社内業務 · 法務・業法","社長・寺井・池之上・弁護士",0,"未着手","言い切り表現・実績表示の法務確認"),
 ("広告","P3","広報・告知（LP・SNS）4件","RC／RCP／HI／PP · REALIZE QUEST","白石",0,"未着手",""),
])]
task_sheet("広告","制作進捗ボード ─ 広告",
  "工程表上のどの項目も0%。制作も運用もチェック体制もまだ始まっていません。",
  "広告を出すには、飛ばす先のLPが公開され、問い合わせが届く状態である必要があります。着手日はLP公開日で決まります。",
  ad)

# ============================================================ 4. 凡例
lg=wb.create_sheet("凡例と出典"); lg.sheet_view.showGridLines=False
lg.column_dimensions["A"].width=24; lg.column_dimensions["B"].width=98
lg["A1"]="凡例と出典"; lg["A1"].font=Font(name=F,size=15,bold=True,color=INK)
def head(r,t):
    c=lg.cell(row=r,column=1,value=t); c.font=Font(name=F,size=12,bold=True,color=ACC)
def kv(r,k,v):
    a=lg.cell(row=r,column=1,value=k); a.font=Font(name=F,size=9,bold=True,color=INK)
    a.alignment=Alignment(vertical="top")
    b=lg.cell(row=r,column=2,value=v); b.font=Font(name=F,size=9,color=INK)
    b.alignment=Alignment(vertical="top",wrap_text=True); lg.row_dimensions[r].height=34
head(3,"LPシートの4段")
kv(4,"①原稿・デザイン","本編の文章とデザインが最後まで通っている")
kv(5,"②付随ページ","資料請求・会社概要など、必要な付随ページが揃っている")
kv(6,"③問い合わせ導線","ボタンの行き先が実在し、問い合わせが届く先につながっている")
kv(7,"④本番公開","独自ドメインで公開され、J列に本番URLが入っている（自動判定）")
head(9,"記号")
kv(10,"○","できている")
kv(11,"△","仮づけ。メーラーを開くだけ、または行き先が「#」のまま")
kv(12,"×","手つかず／ページが存在しない")
head(14,"状態（eラーニング・動画／広告シート）")
kv(15,"確認待ち","作業は終わり、誰かの確認を待っている")
kv(16,"進行中","いま手が動いている")
kv(17,"未着手","まだ始まっていない")
kv(18,"P2／P3","工程表の優先度。P3は「中核が回り出してから戻ってくる」もの")
kv(19,"赤い担当名","担当が「未定」の仕事。誰かを決めないと進みません")
head(21,"使い方")
kv(22,"案件・区分のセル（水色）","LPシートのA列（案件）とB列（状態）、他シートのA列（区分）とB列（優先）は、クリックすると▼が出て選べます。行の並べ替えをしなくても、絞り込みで見たい案件だけ出せます。")
kv(23,"進捗のセル（水色）","クリックすると右に▼が出ます。0%〜100%を5%きざみで選べます（直接入力も可）。選ぶとセルの中のバーがその長さに変わります。")
kv(24,"状態のセル（水色）","クリックして 未着手／進行中／確認待ち を選びます。")
kv(25,"LPの○△×（水色）","E〜G列。クリックして選び直せます。H列の④本番公開はJ列のURL有無から自動で決まります。")
kv(26,"列見出しの▼","各シートの見出し行にオートフィルタが付いています。▼から値を選ぶと、その行だけ表示されます。複数の列を組み合わせられます。")
kv(27,"黄色の記入欄","公開したURL・動画URL・広告の入稿先を貼ります。")
kv(28,"自動で動く数字","LPシートは④と到達度、各シートの最下段の平均・件数が数式です。手で書き換えないでください。")
kv(29,"選択肢を増やしたいとき","非表示シート「選択肢」に一覧が入っています。表示して値を足し、入力規則の参照範囲を広げれば選べるようになります。")
head(31,"出典")
kv(32,"LPシート","公開済みアーティファクト43件のうちLPに該当する22ページ（本編11＋付随11）を、1ページずつHTMLを読んで確認。ボタンの行き先・フォームの送信方法・付随ページの有無は実物ベースです。")
kv(33,"進捗％と担当","社内の「REALIZE OS 制作MASTER」(2026-09-03) と「仕事の棚卸し」(2026-09-02) の数値をそのまま引いています。")
kv(34,"担当の食い違い","広告運用の担当は2資料で記載が違います（制作MASTER＝大熊・片山／棚卸し＝池之上・片山）。広告制作は大熊・片山、広告運用は池之上・片山としています。")
kv(35,"未確認","LP付随ページ4本（RC星空案の会社概要、RCアイス案の資料請求、LMP星空案の資料請求と会社概要）は、親LPからのリンクで存在を確認しただけで中身は開いていません。")
kv(36,"調査日","2026-09-03")
for r in list(range(4,8))+[10,11,12]+list(range(15,20))+list(range(22,30))+list(range(32,37)):
    lg.row_dimensions[r].height=36

opt.sheet_state="hidden"
wb.calculation.fullCalcOnLoad = True
wb.save("/home/user/jigyokeikakusyo/制作進捗ボード.xlsx")
print("saved")
